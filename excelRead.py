# Python program to read an excel file
import pandas as pd
import openpyxl

# Give the location of the file
pathMRI = "MRI_Pelvis_Pedi.F.10yr.xlsx"
pathUS = "US_Pelvis_Pedi.F.10yr.xlsx"

# To open the workbook
# workbook object is created
wb_objMRI = openpyxl.load_workbook(pathMRI)
wb_objUS = openpyxl.load_workbook(pathUS)
ultrasoundMRN = list()
mriMRN = list()


# Get workbook active sheet object from the active attribute
sheet_US = wb_objUS.active
sheet_MRI = wb_objMRI.active

# Getting the value of maximum rows and column
row = sheet_MRI.max_row
column = sheet_MRI.max_column

print("Total Rows:", row) #the rows represent the total number of patients
print("Total Columns:", column)

# Extracting patient MRNs from MRI and Ultrasound data
for i in range(1, row + 1):
	cell_obj = sheet_US.cell(row = i, column = 14)
	ultrasoundMRN.append(cell_obj.value)
	#print(cell_obj.value)

for i in range(1, row + 1):
	cell_obj = sheet_MRI.cell(row = i, column = 14)
	mriMRN.append(cell_obj.value)
	#print(cell_obj.value)

print(ultrasoundMRN)
print(mriMRN)

matchedMRN = list(set(ultrasoundMRN).intersection(mriMRN))
print(matchedMRN)
len(matchedMRN)



# printing the value of first column
# Loop will print all values of first row
# print("\nValue of first row")
# for i in range(1, column + 1):
# 	cell_obj = sheet_obj.cell(row = 2, column = i)
# 	print(cell_obj.value, end = " ")
